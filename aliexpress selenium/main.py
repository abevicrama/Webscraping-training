import openpyxl.workbook
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
import openpyxl
import time



input_text = input('What you want to buy?')

service = Service(executable_path="chromedriver.exe")
driver = webdriver.Chrome(service=service)

#driver.get('https://www.google.com/')
#WebDriverWait(driver, 5).until(
#    EC.presence_of_element_located((By.CLASS_NAME, 'gLFyf'))
#)
#google_input = driver.find_element(By.CLASS_NAME, 'gLFyf')
#google_input.send_keys('aliexpress'+ Keys.ENTER)
#
#WebDriverWait(driver, 5).until(
#    EC.presence_of_element_located((By.PARTIAL_LINK_TEXT,'AliExpress'))
#)
#web_link = driver.find_element(By.PARTIAL_LINK_TEXT,'AliExpress')
#web_link.click()

driver.get('https://best.aliexpress.com/?spm=a2g0o.productlist.logo.1.53b7Kh1nKh1njD&browser_redirect=true')
ali_search_bar = "search--keyword--15P08Ji"

WebDriverWait(driver, 5).until(
    EC.presence_of_element_located((By.CLASS_NAME,ali_search_bar))
)


ali_input = driver.find_element(By.CLASS_NAME,ali_search_bar)
ali_input.send_keys(input_text+ Keys.ENTER)


file  = "details.xlsx"
wb = openpyxl.Workbook(file)
ws = wb.active
wb.save("details.xlsx")
wb.close()

file  = "details.xlsx"
wb = openpyxl.load_workbook(file)
ws = wb.active
headers = ['Product Name', 'Price']
ws.append(headers)
wb.save("details.xlsx")
wb.close()



p=0
def collect_data(driver,p):

    try:
        time.sleep(5)
        
        i=0
        while(i<2):   
            driver.execute_script("window.scrollBy(0,500)","")
            time.sleep(1)
            i+=0.1
        
     
        

        
        driver.find_element(By.TAG_NAME,'body').send_keys(Keys.PAGE_UP)
        driver.find_element(By.TAG_NAME,'body').send_keys(Keys.PAGE_UP)

        WebDriverWait(driver, 15).until(
        EC.presence_of_all_elements_located((By.CLASS_NAME,'multi--titleText--nXeOvyr'))
        )
    
        item_name = driver.find_elements(By.CLASS_NAME,'multi--titleText--nXeOvyr')
        item_price = driver.find_elements(By.CLASS_NAME,'multi--price-sale--U-S0jtj')
        
        item_name_arr=[]
        item_price_arr=[]

        for i in item_name:
            item_name_arr.append(i.get_attribute("innerText"))
        print(item_name_arr)

        for i in item_price:
            item_price_arr.append(i.get_attribute("innerText"))
        print(item_price_arr)
        
        arr_size = len(item_name)

        for j in range(0,arr_size):
            item_arr = []
            item_arr.clear()
            item_arr.append(item_name_arr[j])
            item_arr.append(item_price_arr[j])
            print(item_arr)   
            file  = "details.xlsx"
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            ws.append(item_arr)
            wb.save("details.xlsx")
            wb.close()
        p+=1
        next = driver.find_element(By.CLASS_NAME,'comet-pagination-next')
        next.click()
        time.sleep(5)
        if(p>60):
            driver.quit()
        
        collect_data(driver,p)

    except:
        print("error")
           


collect_data(driver,p)