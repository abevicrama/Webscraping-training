from bs4 import BeautifulSoup
import requests
import json
import time
import openpyxl



response = requests.get('https://api.sarasavi.lk/api/filter-product?isLiteraryFestival=false&isGalleLiterary=false&page=1&publisher=&category=116&stationery=&language=&per_page=40&author=&brand=&min=0&max=&sortType=&sortValue=')

file = "hello.xlsx"
wb = openpyxl.Workbook(file)
ws = wb.active
wb.save("hello.xlsx")

file = "hello.xlsx"
wb = openpyxl.load_workbook(file)
ws = wb.active
headers = ['ISBN','NAME']
ws.append(headers)
wb.save("hello.xlsx")

def collect_data(response):
    
    if response.status_code == 200:  # Check if the request was successful
        data_script = response.json()
        #print(data_script)       
        try:
            for count in range(0,50):
                arr_data = []
                isbn = data_script['data']['data'][count]['isbn']
                name = data_script['data']['data'][count]['name']
                
                arr_data.append(isbn)
                arr_data.append(name)
                file = "hello.xlsx"
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                ws.append(arr_data)
                wb.save("hello.xlsx")

        except:           
            print('Finished - '+ str(data_script['data']['current_page']))           
            try:                
                next_link = data_script['data']['next_page_url']               
                response = requests.get(next_link)
                time.sleep(1)               
                collect_data(response)               
            except:               
                print('finish all pages')
    else:       
        print(f"Failed to fetch data. HTTP Status Code: {response.status_code}")   
collect_data(response)